import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_spreadsheet(filename,column_heading):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    sheet.cell(1,4).value=column_heading
    #to get the coordinate of a cell=sheet.cell(x,y) or sheet[a1] 
    #and cell.value to fetch the value in a cell

    for row in range(2,sheet.max_row+1):
        cell_value = sheet.cell(row,3)
        corrected_price = cell_value.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    values= Reference(sheet, 
               min_row=2,
               max_row=sheet.max_row,
               min_col=4,
               max_col=4 )

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'A13')

    wb.save(filename)


process_spreadsheet('transactions.xlsx','updated_price')
